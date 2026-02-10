import os
import re
import io
import time
import base64
from typing import List, Dict, Set

from dotenv import load_dotenv
from playwright.sync_api import sync_playwright
from openai import OpenAI
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


ICON_RE = re.compile(r"^icon-[a-z][a-z0-9-]*$")  # must start with a letter (excludes icon-130)

from typing import Tuple

def load_row_index_by_name(xlsx_path: str) -> Dict[str, int]:
    """
    Returns {dash_cased_name: row_number} for existing rows.
    """
    if not os.path.exists(xlsx_path):
        return {}

    wb = load_workbook(xlsx_path)
    ws = wb.active

    idx: Dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if name:
            idx[str(name).strip()] = r
    return idx

def load_row_state(xlsx_path: str) -> Dict[str, Dict[str, object]]:
    """
    Returns:
      {
        "icon-name": {"row": 12, "desc": "..." or None}
      }
    """
    if not os.path.exists(xlsx_path):
        return {}

    wb = load_workbook(xlsx_path)
    ws = wb.active

    state: Dict[str, Dict[str, object]] = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        desc = ws.cell(r, 2).value
        if name:
            state[str(name).strip()] = {"row": r, "desc": (str(desc).strip() if desc else "")}
    return state

def upsert_row(ws, row_index_by_name: Dict[str, int], icon_name: str, description: str):
    if icon_name in row_index_by_name:
        r = row_index_by_name[icon_name]
        ws.cell(r, 2).value = description
    else:
        ws.append([icon_name, description])
        row_index_by_name[icon_name] = ws.max_row

def ensure_workbook(xlsx_path: str):
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        ws = wb.active
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = "icons"
    ws.append(["dash_cased_name", "description"])
    return wb, ws


def autosize_columns(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 80)


def extract_icon_names(page) -> List[str]:
    # Pull icon classes from the DOM, not text, and exclude cms-icon-* sizing classes
    classes = page.evaluate("""
      () => {
        const out = new Set();
        document.querySelectorAll('[class*="icon-"]').forEach(el => {
          el.classList.forEach(c => out.add(c));
        });
        return Array.from(out);
      }
    """)

    names = sorted({
        c for c in classes
        if ICON_RE.match(c) and not c.startswith("cms-icon-")
    })
    return names



def build_render_surface(page):
    # Create a fixed render area to screenshot individual icons consistently
    page.add_style_tag(content="""
      #__iconShotHost {
        position: fixed;
        left: 24px;
        top: 24px;
        width: 260px;
        height: 260px;
        background: white;
        display: flex;
        align-items: center;
        justify-content: center;
        z-index: 999999;
        border: 1px solid rgba(0,0,0,0.1);
      }
      #__iconShotHost span {
        display: inline-block;
      }
    """)
    page.evaluate("""
      () => {
        const existing = document.getElementById('__iconShotHost');
        if (existing) existing.remove();
        const host = document.createElement('div');
        host.id = '__iconShotHost';
        document.body.appendChild(host);
      }
    """)


def render_icon(page, icon_class: str):
    # Use cms-icon-200 for legibility; it’s documented as an icon size on the page.  :contentReference[oaicite:0]{index=0}
    page.evaluate(
        """(cls) => {
            const host = document.getElementById('__iconShotHost');
            host.innerHTML = '';
            const el = document.createElement('span');
            el.className = `cms-icon-200 ${cls}`;
            host.appendChild(el);
        }""",
        icon_class
    )
    return page.locator("#__iconShotHost")


def png_bytes_to_data_url(png_bytes: bytes) -> str:
    b64 = base64.b64encode(png_bytes).decode("ascii")
    return f"data:image/png;base64,{b64}"


def describe_icon(client: OpenAI, icon_name: str, png_bytes: bytes) -> str:
    data_url = png_bytes_to_data_url(png_bytes)

    prompt = (
        "You are describing a UI icon for a spreadsheet.\n"
        f"Icon class name: {icon_name}\n\n"
        "Return ONE short, concrete description (5–12 words).\n"
        "Focus on the visible glyph only. No guesses about product context."
    )

    # Responses API supports image inputs via `input_image` with a data URL. :contentReference[oaicite:1]{index=1}
    resp = client.responses.create(
        model="gpt-4.1-mini",
        input=[{
            "role": "user",
            "content": [
                {"type": "input_text", "text": prompt},
                {"type": "input_image", "image_url": data_url},
            ],
        }],
    )
    return (resp.output_text or "").strip()


def main():
    load_dotenv()

    url = os.getenv("ICON_LIST_URL", "https://devnet.kentico.com/docs/icon-list/index.html")
    out_xlsx = os.getenv("OUTPUT_XLSX", "kentico-icons.xlsx")
    shots_dir = os.path.join(os.getcwd(), "icon_shots")
    os.makedirs(shots_dir, exist_ok=True)

    client = OpenAI()

    row_state = load_row_state(out_xlsx)
    row_index_by_name = {k: v["row"] for k, v in row_state.items()}

    wb, ws = ensure_workbook(out_xlsx)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(viewport={"width": 1280, "height": 900})
        page.goto(url, wait_until="networkidle")

        page.wait_for_timeout(250)
        page.evaluate("() => document.fonts && document.fonts.ready")
        page.wait_for_timeout(250)


        icon_names = extract_icon_names(page)
        build_render_surface(page)

        total = len(icon_names)
        start_time = time.time()


        for i, icon_name in enumerate(icon_names, start=1):
            existing_desc = row_state.get(icon_name, {}).get("desc", "")
            if existing_desc and not existing_desc.startswith("[ERROR]"):
                # Already successfully processed; skip entirely
                continue

            png_path = os.path.join(shots_dir, f"{icon_name}.png")

            # 1) Screenshot ONLY if the PNG does not already exist
            if os.path.exists(png_path):
                with open(png_path, "rb") as f:
                    png_bytes = f.read()
            else:
                host = render_icon(page, icon_name)
                host.wait_for(state="visible")
                png_bytes = host.screenshot(type="png")
                with open(png_path, "wb") as f:
                    f.write(png_bytes)

            # 2) Always attempt AI analysis
            try:
                desc = describe_icon(client, icon_name, png_bytes)
            except Exception as e:
                msg = str(e)

                # Stop immediately if quota is exhausted
                if "insufficient_quota" in msg or "exceeded your current quota" in msg:
                    raise SystemExit(
                        "OpenAI quota exhausted (insufficient_quota). "
                        "Fix billing or switch API key, then rerun."
                    )

                desc = f"[ERROR] {type(e).__name__}: {e}"

            # 3) Update or insert spreadsheet row
            upsert_row(ws, row_index_by_name, icon_name, desc)

            autosize_columns(ws)
            wb.save(out_xlsx)

            elapsed = time.time() - start_time
            avg = elapsed / i
            remaining = avg * (total - i)

            print(
                f"\r[{i}/{total}] {icon_name} "
                f"| avg {avg:.2f}s/icon "
                f"| ETA {remaining/60:.1f}m",
                end="",
                flush=True
            )

            time.sleep(0.15)


        browser.close()

    print(f"Done. Wrote: {out_xlsx}")
    print(f"Icon PNGs in: {shots_dir}")


if __name__ == "__main__":
    main()
