#!/usr/bin/env python3
"""
Add XML doc comments to Kentico icon constants based on an XLSX mapping.

- Reads an XLSX with columns:
  A: dash_cased_name (e.g., icon-accordion)
  B: description (e.g., "Accordion / expanding panels")

- For each line like:
    public const string ACCORDION = "icon-accordion";

  If there is a matching icon name in the XLSX and the const does NOT already
  have an XML doc comment immediately above it, insert:

    /// <summary>
    /// Accordion / expanding panels
    /// </summary>

- Keeps #pragma directives in place. If a const is preceded by #pragma lines,
  the doc comment is inserted between the #pragma line(s) and the const (so the
  comment still attaches to the const).

Usage:
  python add_icon_docs.py --cs KenticoIcons.cs --xlsx kentico-icons.xlsx --out KenticoIcons.cs

Optional:
  --inplace     overwrite the input .cs file (ignores --out)
  --sheet icons specify worksheet name (defaults to active sheet)
"""

import argparse
import os
import re
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


CONST_RE = re.compile(
    r'^(?P<indent>\s*)public\s+const\s+string\s+(?P<field>[A-Z0-9_]+)\s*=\s*"(?P<icon>icon-[a-z0-9-]+)"\s*;\s*$'
)
PRAGMA_RE = re.compile(r"^\s*#pragma\b")
XMLDOC_RE = re.compile(r"^\s*///\s*<summary>\s*$")


def xml_escape(s: str) -> str:
    return (
        s.replace("&", "&amp;")
         .replace("<", "&lt;")
         .replace(">", "&gt;")
    )


def load_icon_map(xlsx_path: str, sheet_name: Optional[str] = None) -> Dict[str, str]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    icon_map: Dict[str, str] = {}
    # Expect header row: dash_cased_name | description
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        name = (str(row[0]).strip() if row[0] is not None else "")
        desc = (str(row[1]).strip() if len(row) > 1 and row[1] is not None else "")
        if not name or not desc:
            continue
        # Skip error rows if they exist
        if desc.startswith("[ERROR]"):
            continue
        icon_map[name] = desc
    return icon_map


def last_nonempty_line_idx(lines: List[str]) -> Optional[int]:
    for i in range(len(lines) - 1, -1, -1):
        if lines[i].strip() != "":
            return i
    return None


def already_has_doccomment(out_lines: List[str]) -> bool:
    """
    True if the immediately preceding non-empty line is part of an XML doc comment
    (i.e., starts with ///). This is a conservative check to avoid duplicating docs.
    """
    idx = last_nonempty_line_idx(out_lines)
    if idx is None:
        return False
    return out_lines[idx].lstrip().startswith("///")


def make_doc_block(indent: str, description: str) -> List[str]:
    desc = xml_escape(description)
    return [
        f"{indent}/// <summary>\n",
        f"{indent}/// {desc}\n",
        f"{indent}/// </summary>\n",
    ]


def transform(cs_path: str, xlsx_path: str, out_path: str, sheet: Optional[str]) -> Tuple[int, int]:
    icon_map = load_icon_map(xlsx_path, sheet_name=sheet)

    with open(cs_path, "r", encoding="utf-8") as f:
        in_lines = f.readlines()

    out_lines: List[str] = []
    inserted = 0
    matched = 0

    for line in in_lines:
        m = CONST_RE.match(line)
        if not m:
            out_lines.append(line)
            continue

        icon = m.group("icon")
        indent = m.group("indent")

        if icon not in icon_map:
            out_lines.append(line)
            continue

        matched += 1

        # If the const already has a doc comment immediately above, don't add another
        if already_has_doccomment(out_lines):
            out_lines.append(line)
            continue

        # Special case: if previous line(s) are #pragma directives, doc must go AFTER them
        # (so it still attaches to the const). We only need to handle the immediate
        # previous contiguous pragma lines in the output.
        # We'll pop trailing pragma lines, insert doc, then re-add pragma lines.
        pragma_tail: List[str] = []
        while out_lines:
            prev = out_lines[-1]
            if PRAGMA_RE.match(prev) and (len(pragma_tail) == 0 or PRAGMA_RE.match(prev)):
                pragma_tail.append(out_lines.pop())
                continue
            break
        pragma_tail.reverse()

        # Put pragma lines back first, then doc, then const? No: doc must be directly above const,
        # so pragma must stay above doc. That means we output pragma lines, then doc, then const.
        out_lines.extend(pragma_tail)
        out_lines.extend(make_doc_block(indent, icon_map[icon]))
        out_lines.append(line)
        inserted += 1

    with open(out_path, "w", encoding="utf-8", newline="") as f:
        f.writelines(out_lines)

    return matched, inserted


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--cs", required=True, help="Path to the C# file (KenticoIcons.cs)")
    ap.add_argument("--xlsx", required=True, help="Path to the XLSX (kentico-icons.xlsx)")
    ap.add_argument("--out", help="Output .cs path (ignored if --inplace)")
    ap.add_argument("--inplace", action="store_true", help="Overwrite the input .cs file")
    ap.add_argument("--sheet", default=None, help="Worksheet name (default: active sheet)")
    args = ap.parse_args()

    cs_path = args.cs
    xlsx_path = args.xlsx

    if not os.path.exists(cs_path):
        raise SystemExit(f"Missing C# file: {cs_path}")
    if not os.path.exists(xlsx_path):
        raise SystemExit(f"Missing XLSX file: {xlsx_path}")

    out_path = cs_path if args.inplace else (args.out or (cs_path + ".withdocs.cs"))

    matched, inserted = transform(cs_path, xlsx_path, out_path, args.sheet)
    print(f"Icons in XLSX matched in C#: {matched}")
    print(f"Doc comments inserted:      {inserted}")
    print(f"Wrote: {out_path}")


if __name__ == "__main__":
    main()
